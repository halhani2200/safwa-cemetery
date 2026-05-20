"""Read and analyze Excel file of deceased people"""
import sys
import json
import io
import openpyxl
from pathlib import Path

# Force UTF-8 stdout on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_PATH = r"C:\Users\Hussain\Downloads\متوفين 2023.xlsx"
OUTPUT_JSON = r"C:\Users\Hussain\Downloads\safwa-cemetery-site\scripts\excel_analysis.json"

def main():
    if not Path(EXCEL_PATH).exists():
        print(f"ERROR: File not found: {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    analysis = {
        "sheets": [],
        "total_records": 0
    }

    print("=" * 70)
    print("Excel Analysis")
    print("=" * 70)
    print(f"Sheet names: {wb.sheetnames}")
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            "name": sheet_name,
            "dimensions": f"{ws.max_row} rows x {ws.max_column} cols",
            "headers": [],
            "sample_rows": [],
            "non_empty_rows": 0
        }

        print(f"\n--- Sheet: {sheet_name} ---")
        print(f"Dimensions: {ws.max_row} rows x {ws.max_column} cols")

        # Headers
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            headers.append(cell.value)
        sheet_data["headers"] = headers

        # Sample rows
        for row_idx in range(2, min(12, ws.max_row + 1)):
            row_data = {}
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                header = headers[col_idx - 1] if col_idx <= len(headers) else f"col_{col_idx}"
                value = cell.value
                if value is not None:
                    row_data[str(header)] = str(value) if not isinstance(value, (int, float, bool)) else value
            if row_data:
                sheet_data["sample_rows"].append(row_data)

        # Count non-empty rows
        non_empty = 0
        for row_idx in range(2, ws.max_row + 1):
            if any(ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)):
                non_empty += 1
        sheet_data["non_empty_rows"] = non_empty
        analysis["total_records"] += non_empty

        analysis["sheets"].append(sheet_data)

    # Write JSON
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2)

    print(f"\nTotal records across all sheets: {analysis['total_records']}")
    print(f"\nFull analysis written to: {OUTPUT_JSON}")

if __name__ == "__main__":
    main()
