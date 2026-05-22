# -*- coding: utf-8 -*-
"""Analyze 2024 deceased Excel — full A/B picture (UTF-8 safe)."""
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

PATH = r"C:\Users\Hussain\Downloads\deceased_2024.xlsx"
wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["Sheet1"]

A, B = [], []
for r in range(4, ws.max_row + 1):
    name = ws.cell(row=r, column=4).value
    if not name:
        continue
    rec = {
        "xr": r,
        "g1": ws.cell(row=r, column=1).value,           # global cumulative number
        "sec": (str(ws.cell(row=r, column=13).value).strip() if ws.cell(row=r, column=13).value else ""),
        "row": ws.cell(row=r, column=14).value,
        "num": ws.cell(row=r, column=15).value,         # number (A=global, B=local)
        "ref": ws.cell(row=r, column=16).value,
        "name": str(name).strip(),
    }
    if rec["sec"] == "أ":
        A.append(rec)
    elif rec["sec"] == "ب":
        B.append(rec)

ag = [int(a["g1"]) for a in A if isinstance(a["g1"], (int, float))]
print("AREA A: count", len(A), "| g1 range", min(ag), "-", max(ag))
print("AREA A entries with g1>187:",
      [(a["g1"], a["row"], a["ref"], a["name"][:16]) for a in A if isinstance(a["g1"], (int, float)) and a["g1"] > 187])
print("AREA A missing in 119..187:", sorted(set(range(119, 188)) - set(ag)))

print("\nAREA B: count", len(B))
seen = {}
for b in B:
    dup = " <DUP>" if b["ref"] in seen else ""
    seen[b["ref"]] = seen.get(b["ref"], 0) + 1
    print(f"  g1={b['g1']} row={b['row']} num={b['num']} ref={b['ref']} | {b['name'][:22]}{dup}")
