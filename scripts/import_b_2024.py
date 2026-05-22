# -*- coding: utf-8 -*-
"""Generate INSERT SQL for Area B (المنطقة ب) 2024 deceased. Does NOT touch the DB."""
import openpyxl, sys, io, math, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

PATH = r"C:\Users\Hussain\Downloads\deceased_2024.xlsx"
OUT = r"C:\Users\Hussain\Downloads\safwa-cemetery-site\scripts\import_b.sql"

# Area B calibration from user's Google Maps points (right=east=first grave)
E1 = (26.649304, 49.959176)  # row 1 first (east/right)
W1 = (26.649269, 49.958928)  # row 1 last  (west/left)
E2 = (26.649270, 49.959284)  # row 2 first (east/right)
W2 = (26.649238, 49.958935)  # row 2 west extent (physical end)
ROW1_COUNT = 13              # row 1 is full with 13 graves


def hav(a, b):
    R = 6371000.0
    la1, lo1, la2, lo2 = map(math.radians, [a[0], a[1], b[0], b[1]])
    h = math.sin((la2 - la1) / 2) ** 2 + math.cos(la1) * math.cos(la2) * math.sin((lo2 - lo1) / 2) ** 2
    return 2 * R * math.asin(math.sqrt(h))


def lerp(a, b, f):
    return (a[0] + (b[0] - a[0]) * f, a[1] + (b[1] - a[1]) * f)


S = hav(E1, W1) / (ROW1_COUNT - 1)   # uniform spacing derived from full row 1
D2 = hav(E2, W2)


def coord(row, pos):                  # pos = 1-based within its row (1 = east/first)
    if row == 1:
        return lerp(E1, W1, (pos - 1) / (ROW1_COUNT - 1))
    return lerp(E2, W2, (pos - 1) * S / D2)   # row 2: fixed spacing from east end


def norm(t):
    if not t:
        return ""
    t = str(t)
    t = re.sub(r'[ً-ْٰ]', '', t)
    t = re.sub(r'[إأآا]', 'ا', t)
    for a, b in [('ى', 'ي'), ('ة', 'ه'), ('ؤ', 'و'), ('ئ', 'ي'), ('ـ', '')]:
        t = t.replace(a, b)
    t = re.sub(r'\bآل\s*', '', t)
    t = re.sub(r'\bال', '', t)
    return re.sub(r'\s+', ' ', t).strip()


def esc(v):
    if v is None or v == '':
        return 'NULL'
    if isinstance(v, (int, float)):
        return str(v)
    return "'" + str(v).replace("'", "''") + "'"


def fix_gender(g):
    if not g:
        return None
    g = str(g).strip()
    if 'نث' in g:
        return 'انثى'
    if 'ذكر' in g or 'ذكور' in g:
        return 'ذكر'
    return None


wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb["Sheet1"]
rows, seen = [], set()
for r in range(4, ws.max_row + 1):
    name = ws.cell(row=r, column=4).value
    if not name:
        continue
    sec = ws.cell(row=r, column=13).value
    if (str(sec).strip() if sec is not None else "") != "ب":
        continue
    num = ws.cell(row=r, column=15).value
    if num in seen:           # dedupe duplicate row
        continue
    seen.add(num)
    rows.append({
        "name": str(name).strip(),
        "gender": fix_gender(ws.cell(row=r, column=5).value),
        "age": ws.cell(row=r, column=6).value,
        "receiver": ws.cell(row=r, column=7).value,
        "rel": ws.cell(row=r, column=8).value,
        "phone": ws.cell(row=r, column=9).value,
        "day": ws.cell(row=r, column=10).value,
        "hijri": ws.cell(row=r, column=11).value,
        "greg": ws.cell(row=r, column=12).value,
        "row": int(ws.cell(row=r, column=14).value),
        "num": int(num),
        "ref": str(ws.cell(row=r, column=16).value).replace('_', '-'),
        "report": ws.cell(row=r, column=17).value,
    })

rows.sort(key=lambda x: x["num"])
out = ["-- Import Area B (المنطقة ب) 2024 deceased — auto-generated, review before applying"]
serial = 187
for rec in rows:
    serial += 1
    pos = rec["num"] if rec["row"] == 1 else rec["num"] - ROW1_COUNT
    lat, lng = coord(rec["row"], pos)
    lat, lng = round(lat, 6), round(lng, 6)
    out.append(
        "INSERT INTO graves (serial_number,name,normalized_name,gender,age,death_day,"
        "death_date_hijri,death_date_gregorian,section,row_number,grave_number,grave_reference,"
        "latitude,longitude,receiver_name,receiver_relationship,receiver_phone,report_number) VALUES ("
        f"{serial},{esc(rec['name'])},{esc(norm(rec['name']))},{esc(rec['gender'])},{esc(rec['age'])},"
        f"{esc(rec['day'])},{esc(rec['hijri'])},{esc(rec['greg'])},'ب',{rec['row']},{rec['num']},"
        f"{esc(rec['ref'])},{lat},{lng},{esc(rec['receiver'])},{esc(rec['rel'])},{esc(rec['phone'])},"
        f"{esc(str(rec['report']) if rec['report'] else None)});"
    )

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(out))

print(f"spacing S = {S:.3f} m  |  row2 length D2 = {D2:.3f} m")
print(f"unique Area B graves to insert: {len(rows)}  (serial 188..{serial})")
print("verification (corner graves):")
for rec in rows:
    if rec["num"] in (1, 13, 14, 23):
        pos = rec["num"] if rec["row"] == 1 else rec["num"] - ROW1_COUNT
        lat, lng = coord(rec["row"], pos)
        print(f"  ب-{rec['row']}-{rec['num']} pos{pos}: {round(lat,6)}, {round(lng,6)}  | {rec['name'][:20]}")
print("OUT:", OUT)
