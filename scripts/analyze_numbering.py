"""Check numbering pattern in Excel"""
import sys
import io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_PATH = r"C:\Users\Hussain\Downloads\متوفين 2023.xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["Sheet1"]

print("Row | الصف | الرقم | رقم القبر | الاسم (first 30 chars)")
print("-" * 80)
prev_row = None
for row_idx in range(4, min(40, ws.max_row + 1)):
    row_num = ws.cell(row=row_idx, column=12).value
    grave_num = ws.cell(row=row_idx, column=13).value
    grave_ref = ws.cell(row=row_idx, column=14).value
    name = ws.cell(row=row_idx, column=2).value
    if not name:
        continue
    marker = ""
    if prev_row is not None and row_num != prev_row:
        marker = " <-- NEW ROW"
    print(f"{row_idx} | {row_num} | {grave_num} | {grave_ref} | {(name or '')[:30]}{marker}")
    prev_row = row_num
