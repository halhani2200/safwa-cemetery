"""Better Excel analysis - shows raw structure"""
import sys
import json
import io
import openpyxl
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_PATH = r"C:\Users\Hussain\Downloads\متوفين 2023.xlsx"
OUTPUT_FILE = r"C:\Users\Hussain\Downloads\safwa-cemetery-site\scripts\excel_full_dump.txt"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["Sheet1"]

print(f"Sheet1 dimensions: {ws.max_row} rows x {ws.max_column} cols")

# Find which columns actually have data
print("\nAnalyzing which columns have data...")
col_data_count = {}
for col_idx in range(1, ws.max_column + 1):
    count = sum(1 for row_idx in range(1, ws.max_row + 1) if ws.cell(row=row_idx, column=col_idx).value is not None)
    if count > 0:
        col_data_count[col_idx] = count
        print(f"  Column {col_idx}: {count} non-null cells")

# Dump first 20 rows to file
with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    f.write(f"Sheet1 - dimensions: {ws.max_row} rows x {ws.max_column} cols\n")
    f.write(f"Columns with data: {list(col_data_count.keys())}\n\n")

    # Write first 20 rows showing all non-null columns
    for row_idx in range(1, min(21, ws.max_row + 1)):
        f.write(f"\n=== Row {row_idx} ===\n")
        for col_idx in col_data_count.keys():
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                f.write(f"  Col{col_idx}: {repr(cell.value)}\n")

    # Write rows 270-293 (last few)
    f.write(f"\n\n=== LAST FEW ROWS ===\n")
    for row_idx in range(max(1, ws.max_row - 20), ws.max_row + 1):
        f.write(f"\n=== Row {row_idx} ===\n")
        for col_idx in col_data_count.keys():
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                f.write(f"  Col{col_idx}: {repr(cell.value)}\n")

print(f"\nFull dump written to: {OUTPUT_FILE}")
