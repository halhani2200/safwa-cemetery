"""Convert Excel deceased data to SQL INSERT statements"""
import sys
import io
import openpyxl
from pathlib import Path
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_PATH = r"C:\Users\Hussain\Downloads\متوفين 2023.xlsx"
OUTPUT_SQL = r"C:\Users\Hussain\Downloads\safwa-cemetery-site\scripts\import_data.sql"

# Section A calibration points (from earlier work)
# Bilinear interpolation corners for Section A
SECTION_A_CORNERS = {
    "nw": (26.649576, 49.958884),  # Row 1, Grave 1 (NW corner)
    "ne": (26.649639, 49.959215),  # Row 1, Grave 17 (NE corner)
    "sw": (26.649334, 49.958942),  # Row 11, Grave 1 (SW corner)
    "se": (26.649388, 49.959267)   # Row 11, Grave 17 (SE corner)
}

ROWS_A = 11
COLS_A = 17

def calc_coordinates(section, row, grave_number):
    """Calculate coordinates using bilinear interpolation for Section A"""
    if section != "أ" or not row or not grave_number:
        return None, None

    try:
        row = int(row)
        grave_number = int(grave_number)
    except (TypeError, ValueError):
        return None, None

    if row < 1 or row > ROWS_A or grave_number < 1 or grave_number > COLS_A:
        return None, None

    u = (grave_number - 1) / (COLS_A - 1)  # column fraction 0-1
    v = (row - 1) / (ROWS_A - 1)            # row fraction 0-1

    nw_lat, nw_lng = SECTION_A_CORNERS["nw"]
    ne_lat, ne_lng = SECTION_A_CORNERS["ne"]
    sw_lat, sw_lng = SECTION_A_CORNERS["sw"]
    se_lat, se_lng = SECTION_A_CORNERS["se"]

    lat = (1 - u) * (1 - v) * nw_lat + u * (1 - v) * ne_lat + (1 - u) * v * sw_lat + u * v * se_lat
    lng = (1 - u) * (1 - v) * nw_lng + u * (1 - v) * ne_lng + (1 - u) * v * sw_lng + u * v * se_lng

    return round(lat, 6), round(lng, 6)


def normalize_arabic(text):
    """Normalize Arabic text for search (same rules as in JS)"""
    if not text:
        return ""
    text = str(text)
    # Remove diacritics
    text = re.sub(r'[ً-ٰٟ]', '', text)
    # Normalize alif variations
    text = re.sub(r'[إأآا]', 'ا', text)
    # Normalize yaa
    text = text.replace('ى', 'ي')
    # Normalize taa marbuta
    text = text.replace('ة', 'ه')
    # Normalize hamza on waw/yaa
    text = text.replace('ؤ', 'و')
    text = text.replace('ئ', 'ي')
    # Remove tatweel
    text = text.replace('ـ', '')
    # Normalize "آل" prefix (remove)
    text = re.sub(r'\bآل\s*', '', text)
    # Normalize "ال" prefix (remove)
    text = re.sub(r'\bال', '', text)
    # Multiple spaces to one
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def sql_escape(value):
    """Escape value for SQL"""
    if value is None or value == '':
        return 'NULL'
    if isinstance(value, (int, float)):
        return str(value)
    s = str(value)
    s = s.replace("'", "''")
    return f"'{s}'"


def clean_grave_ref(ref):
    """Standardize grave reference like 'أ-1-1' (replace underscores)"""
    if not ref:
        return None
    s = str(ref).strip()
    s = s.replace('_', '-')
    return s


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Sheet1"]

    insert_statements = []
    insert_statements.append("-- Import 2023 deceased data")
    insert_statements.append("-- Auto-generated from Excel\n")

    count = 0
    skipped = 0

    # Data starts at row 4 (rows 1-2 empty, row 3 is header)
    for row_idx in range(4, ws.max_row + 1):
        serial = ws.cell(row=row_idx, column=1).value
        name = ws.cell(row=row_idx, column=2).value

        # Skip empty rows
        if not name:
            skipped += 1
            continue

        gender = ws.cell(row=row_idx, column=3).value
        age = ws.cell(row=row_idx, column=4).value
        receiver_name = ws.cell(row=row_idx, column=5).value
        receiver_relationship = ws.cell(row=row_idx, column=6).value
        receiver_phone = ws.cell(row=row_idx, column=7).value
        death_day = ws.cell(row=row_idx, column=8).value
        death_date_hijri = ws.cell(row=row_idx, column=9).value
        death_date_gregorian = ws.cell(row=row_idx, column=10).value
        section = ws.cell(row=row_idx, column=11).value
        if not section:
            section = 'أ'  # Default to Section A for 2023 data
        row_num = ws.cell(row=row_idx, column=12).value
        grave_num = ws.cell(row=row_idx, column=13).value
        grave_ref = clean_grave_ref(ws.cell(row=row_idx, column=14).value)
        report_number = ws.cell(row=row_idx, column=15).value

        # Calculate coordinates
        lat, lng = calc_coordinates(section, row_num, grave_num)

        # Normalize name for search
        normalized = normalize_arabic(str(name).strip())

        # Build INSERT statement
        sql = f"""INSERT INTO graves (
    serial_number, name, normalized_name, gender, age,
    death_day, death_date_hijri, death_date_gregorian,
    section, row_number, grave_number, grave_reference,
    latitude, longitude,
    receiver_name, receiver_relationship, receiver_phone, report_number
) VALUES (
    {sql_escape(serial)},
    {sql_escape(str(name).strip())},
    {sql_escape(normalized)},
    {sql_escape(gender)},
    {sql_escape(age)},
    {sql_escape(death_day)},
    {sql_escape(death_date_hijri)},
    {sql_escape(death_date_gregorian)},
    {sql_escape(section)},
    {sql_escape(row_num)},
    {sql_escape(grave_num)},
    {sql_escape(grave_ref)},
    {sql_escape(lat)},
    {sql_escape(lng)},
    {sql_escape(receiver_name)},
    {sql_escape(receiver_relationship)},
    {sql_escape(receiver_phone)},
    {sql_escape(str(report_number) if report_number else None)}
);"""
        insert_statements.append(sql)
        count += 1


    with open(OUTPUT_SQL, "w", encoding="utf-8") as f:
        f.write("\n".join(insert_statements))

    print(f"Generated {count} INSERT statements (skipped {skipped} empty rows)")
    print(f"Output: {OUTPUT_SQL}")


if __name__ == "__main__":
    main()
